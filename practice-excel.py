from openpyxl import load_workbook
#load_workbook 只能打开已经存在的 Excel，不能创建新的工作簿。
work_book=load_workbook(filename=r'C:\Users\13691\Desktop\excel\mydjango_course.xlsx')
#载入excel，输出表名
print(work_book.sheetnames)
#根据表名称获取工作表
sheet=work_book['mydjango_course']
#获取工作表的数据范围
print(sheet.dimensions)
# 获取某个单元格的具体内容
#指定行列数
cell=sheet.cell(row=2,column=2)
print(cell.value)
#获取某行，用数字表示，范围[1:3],输出值需要遍历
cells=sheet[2]
for cell in cells:
    print(cell.value)
#获取某列，用字母表示，范围[A:E]
cells=sheet['B']
for cell in cells:
    print(cell.value)

#读取所有行，即整个表格
for row in sheet.rows:
    for cell in row:
        print(cell.value)

