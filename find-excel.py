from openpyxl import load_workbook, Workbook
import glob

goal_dir = r'C:\Users\13691\Desktop\excel'
new_workbook = Workbook()
new_sheet = new_workbook.active

# 用 flag 变量明确新表是否已经添加了表头，只要添加过一次就无须重复再添加
flag = 0

for file in glob.glob(goal_dir + '/*.xlsx'):
    workbook = load_workbook(file)
    sheet = workbook.active

    stu_num = sheet['E']
    row_lst = []
    for cell in stu_num:
        if isinstance(cell.value, int) and cell.value >= 90:
            print(cell.row)
            row_lst.append(cell.row)

    if not flag:
        header = sheet[1]
        header_lst = []
        for cell in header:
            header_lst.append(cell.value)
        new_sheet.append(header_lst)
        flag = 1

    for row in row_lst:
        data_lst = []
        for cell in sheet[row]:
            data_lst.append(cell.value)
        new_sheet.append(data_lst)

new_workbook.save(goal_dir + '/' + 'new_baby_trade.xlsx')